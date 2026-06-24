from __future__ import annotations

import argparse
import csv
import inspect
import json
import os
import statistics
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = REPO_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

from benchmark_danmaku_sequence import (  # type: ignore
    ModelContext,
    append_jsonl,
    capture_timestamp,
    collect_images,
)
from benchmark_model_latency import percentile  # type: ignore
from danmaku.api.llm_client import LLMClient  # type: ignore
from danmaku.models import CaptureFrame, CommentBatch  # type: ignore


DEFAULT_MODEL = "gemini-3.1-flash-lite"


@dataclass(slots=True)
class VariantConfig:
    name: str
    current_dimension: int
    current_quality: int
    use_multi_frame: bool
    history_count: int
    history_dimension: int
    history_quality: int


@dataclass(slots=True)
class MultiFrameResult:
    logged_at: str
    variant: str
    use_multi_frame: bool
    history_count_requested: int
    history_count_sent: int
    current_dimension: int
    current_quality: int
    history_dimension: int
    history_quality: int
    frame_index: int
    capture_timestamp: float
    image_path: str
    source_bytes: int
    model: str
    latency_sec: float
    image_preparation_sec: float
    api_duration_sec: float
    response_parsing_sec: float
    end_to_end_sec: float
    first_attempt_sec: float
    retry_used: bool
    retry_duration_sec: float | None
    ok: bool
    comments: list[str]
    long_comments: list[str]
    summary: str
    scene_change_detected: bool
    context_sent: str
    recent_comments_sent: list[str]
    summary_history: list[str]
    recent_comment_history: list[str]
    consecutive_api_failures: int
    error_message: str
    historical_paths: list[str]


def dimension_name(max_dimension: int) -> str:
    return "original" if max_dimension <= 0 else str(max_dimension)


def build_capture_frame(image_path: Path, frame_index: int) -> CaptureFrame:
    return CaptureFrame(
        image_path=image_path,
        timestamp=capture_timestamp(image_path, frame_index),
        ocr_text="",
    )


def build_client(
    model: str,
    current_dimension: int,
    current_quality: int,
    max_output_tokens: int,
    history_dimension: int,
    history_quality: int,
) -> LLMClient | None:
    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        print("[skip] missing GEMINI_API_KEY")
        return None

    kwargs: dict[str, Any] = {
        "api_key": api_key,
        "api_provider": "gemini",
        "model_name": model,
        "use_dummy_api": False,
        "send_screenshot": True,
        "image_max_dimension": current_dimension,
        "image_jpeg_quality": current_quality,
        "max_output_tokens": max_output_tokens,
        "save_api_images": False,
    }

    params = inspect.signature(LLMClient).parameters
    history_dimension_keys = [
        "historical_frame_max_dimension",
        "history_frame_max_dimension",
        "historical_image_max_dimension",
        "history_image_max_dimension",
        "api_history_image_max_dimension",
    ]
    history_quality_keys = [
        "historical_frame_jpeg_quality",
        "history_frame_jpeg_quality",
        "historical_image_jpeg_quality",
        "history_image_jpeg_quality",
        "api_history_image_jpeg_quality",
    ]

    for key in history_dimension_keys:
        if key in params:
            kwargs[key] = history_dimension
            break
    for key in history_quality_keys:
        if key in params:
            kwargs[key] = history_quality
            break

    return LLMClient(**kwargs)


def add_metrics(
    totals: dict[str, float],
    metrics: dict[str, float],
) -> None:
    for name, value in metrics.items():
        totals[name] = totals.get(name, 0.0) + float(value)


def call_model(
    client: LLMClient,
    frame: CaptureFrame,
    context_sent: str,
    recent_comments_sent: list[str],
    historical_frames: list[CaptureFrame] | None,
) -> tuple[CommentBatch, bool]:
    kwargs: dict[str, Any] = {
        "frame": frame,
        "previous_summary": context_sent,
        "previous_comments": recent_comments_sent,
        "use_streaming": False,
    }

    history_arg_used = False
    if historical_frames:
        generate_params = inspect.signature(client.generate_comments).parameters
        for key in (
            "historical_frames",
            "history_frames",
            "previous_frames",
            "recent_frames",
        ):
            if key in generate_params:
                kwargs[key] = historical_frames
                history_arg_used = True
                break

    batch = client.generate_comments(**kwargs)
    return batch, history_arg_used


def history_paths_for_index(
    images: list[Path],
    frame_index: int,
    history_count: int,
) -> list[Path]:
    if history_count <= 0 or frame_index <= 1:
        return []
    start = max(0, frame_index - 1 - history_count)
    end = frame_index - 1
    return images[start:end]


def generate_frame(
    client: LLMClient,
    model: str,
    config: VariantConfig,
    image_path: Path,
    frame_index: int,
    images: list[Path],
    context: ModelContext,
    retry_count: int,
) -> MultiFrameResult:
    frame = build_capture_frame(image_path, frame_index)
    historical_paths = history_paths_for_index(
        images=images,
        frame_index=frame_index,
        history_count=config.history_count if config.use_multi_frame else 0,
    )
    historical_frames = [
        build_capture_frame(path, idx + 1)
        for idx, path in enumerate(historical_paths)
    ]

    context_sent = context.build_context_summary()
    recent_comments_sent = context.recent_comment_history[-12:]

    overall_started = time.perf_counter()
    timing_totals: dict[str, float] = {}

    first_started = time.perf_counter()
    batch, history_arg_used = call_model(
        client=client,
        frame=frame,
        context_sent=context_sent,
        recent_comments_sent=recent_comments_sent,
        historical_frames=historical_frames if config.use_multi_frame else None,
    )
    first_attempt_sec = round(time.perf_counter() - first_started, 3)
    add_metrics(timing_totals, getattr(client, "last_call_metrics", {}))

    retry_used = False
    retry_duration_sec: float | None = None
    retry_started_total: float | None = None

    for _ in range(retry_count):
        if not batch.is_error:
            break
        retry_used = True
        if retry_started_total is None:
            retry_started_total = time.perf_counter()

        batch, history_arg_used = call_model(
            client=client,
            frame=frame,
            context_sent=context_sent,
            recent_comments_sent=recent_comments_sent,
            historical_frames=(
                historical_frames if config.use_multi_frame else None
            ),
        )
        add_metrics(timing_totals, getattr(client, "last_call_metrics", {}))

    if retry_started_total is not None:
        retry_duration_sec = round(
            time.perf_counter() - retry_started_total,
            3,
        )

    latency_sec = round(time.perf_counter() - overall_started, 3)
    scene_changed = context.apply_batch(batch)

    sent_history_count = len(historical_frames) if history_arg_used else 0

    return MultiFrameResult(
        logged_at=datetime.now().isoformat(timespec="seconds"),
        variant=config.name,
        use_multi_frame=config.use_multi_frame,
        history_count_requested=config.history_count if config.use_multi_frame else 0,
        history_count_sent=sent_history_count,
        current_dimension=config.current_dimension,
        current_quality=config.current_quality,
        history_dimension=config.history_dimension if config.use_multi_frame else 0,
        history_quality=config.history_quality if config.use_multi_frame else 0,
        frame_index=frame_index,
        capture_timestamp=frame.timestamp,
        image_path=str(image_path),
        source_bytes=image_path.stat().st_size,
        model=model,
        latency_sec=latency_sec,
        image_preparation_sec=round(
            timing_totals.get("image_preparation_sec", 0.0),
            6,
        ),
        api_duration_sec=round(
            timing_totals.get("api_duration_sec", 0.0),
            6,
        ),
        response_parsing_sec=round(
            timing_totals.get("response_parsing_sec", 0.0),
            6,
        ),
        end_to_end_sec=round(
            timing_totals.get("end_to_end_sec", latency_sec),
            6,
        ),
        first_attempt_sec=first_attempt_sec,
        retry_used=retry_used,
        retry_duration_sec=retry_duration_sec,
        ok=not batch.is_error,
        comments=batch.comments,
        long_comments=batch.long_comments,
        summary=batch.summary,
        scene_change_detected=scene_changed,
        context_sent=context_sent,
        recent_comments_sent=recent_comments_sent,
        summary_history=context.summary_history[-2:],
        recent_comment_history=context.recent_comment_history[-12:],
        consecutive_api_failures=context.consecutive_api_failures,
        error_message=batch.error_message,
        historical_paths=[str(path) for path in historical_paths],
    )


def write_variant_log(
    result: MultiFrameResult,
    variant_dir: Path,
    max_output_tokens: int,
) -> None:
    record = asdict(result)
    record.update(
        {
            "ocr_text": "",
            "is_error": not result.ok,
            "api_provider": "gemini",
            "send_screenshot_to_api": True,
            "api_max_output_tokens": max_output_tokens,
            "use_streaming_api": False,
            "timing": {
                "image_preparation_sec": result.image_preparation_sec,
                "api_duration_sec": result.api_duration_sec,
                "response_parsing_sec": result.response_parsing_sec,
                "end_to_end_sec": result.end_to_end_sec,
                "first_attempt_sec": result.first_attempt_sec,
                "retry_used": result.retry_used,
                "retry_duration_sec": result.retry_duration_sec,
            },
        }
    )
    append_jsonl(variant_dir / "comments.jsonl", record)


def comparison_cell(result: MultiFrameResult | None) -> str:
    if result is None:
        return "NO RESULT"
    if not result.ok:
        return f"ERROR\n{result.error_message}"

    parts = [
        "TIMING\n"
        f"full response: {result.latency_sec}s\n"
        f"image preparation: {result.image_preparation_sec}s\n"
        f"API duration: {result.api_duration_sec}s\n"
        f"end-to-end: {result.end_to_end_sec}s",
        "FRAME INFO\n"
        f"history requested: {result.history_count_requested}\n"
        f"history sent: {result.history_count_sent}\n"
        f"history files: {len(result.historical_paths)}",
    ]
    if result.comments:
        parts.append(
            "COMMENTS\n" + "\n".join(f"- {item}" for item in result.comments)
        )
    if result.long_comments:
        parts.append(
            "LONG COMMENTS\n"
            + "\n".join(f"- {item}" for item in result.long_comments)
        )
    parts.append(f"SUMMARY\n{result.summary or '(empty)'}")
    return "\n\n".join(parts)


def set_excel_style(ws) -> None:
    header_font = Font(bold=True)
    top_alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = top_alignment

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = top_alignment

    width_map = {
        1: 10,
        2: 40,
        3: 15,
        4: 70,
        5: 70,
    }
    for index, width in width_map.items():
        ws.column_dimensions[get_column_letter(index)].width = width


def write_outputs(
    results: list[MultiFrameResult],
    output_dir: Path,
    image_count: int,
    max_output_tokens: int,
) -> None:
    frame_results_path = output_dir / "frame_results.csv"
    comparison_path = output_dir / "comment_comparison.csv"
    quality_review_path = output_dir / "quality_review.csv"
    summary_path = output_dir / "run_summary.csv"
    workbook_path = output_dir / "comment_quality_comparison.xlsx"

    fields = list(asdict(results[0]).keys())
    with frame_results_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for result in results:
            row = asdict(result)
            for field in (
                "comments",
                "long_comments",
                "recent_comments_sent",
                "summary_history",
                "recent_comment_history",
                "historical_paths",
            ):
                row[field] = json.dumps(row[field], ensure_ascii=False)
            writer.writerow(row)

    variants = list(dict.fromkeys(result.variant for result in results))
    by_frame: dict[int, dict[str, MultiFrameResult]] = {}
    for result in results:
        by_frame.setdefault(result.frame_index, {})[result.variant] = result

    comparison_fields = ["frame_index", "image_path", *variants]
    comparison_rows: list[dict[str, object]] = []
    for frame_index in sorted(by_frame):
        per_variant = by_frame[frame_index]
        first_result = next(iter(per_variant.values()))
        row: dict[str, object] = {
            "frame_index": frame_index,
            "image_path": first_result.image_path,
        }
        for variant in variants:
            row[variant] = comparison_cell(per_variant.get(variant))
        comparison_rows.append(row)

    with comparison_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=comparison_fields)
        writer.writeheader()
        writer.writerows(comparison_rows)

    with quality_review_path.open("w", newline="", encoding="utf-8-sig") as file:
        fields = [
            "frame_index",
            "image_path",
            "single_score",
            "multi_score",
            "winner",
            "notes",
        ]
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for row in comparison_rows:
            writer.writerow(
                {
                    "frame_index": row["frame_index"],
                    "image_path": row["image_path"],
                    "single_score": "",
                    "multi_score": "",
                    "winner": "",
                    "notes": "",
                }
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "comparison"
    ws.append(comparison_fields)
    for row in comparison_rows:
        ws.append([row.get(field, "") for field in comparison_fields])
    set_excel_style(ws)

    ws2 = wb.create_sheet("quality_review")
    review_fields = [
        "frame_index",
        "image_path",
        "single_score",
        "multi_score",
        "winner",
        "notes",
    ]
    ws2.append(review_fields)
    for row in comparison_rows:
        ws2.append(
            [
                row["frame_index"],
                row["image_path"],
                "",
                "",
                "",
                "",
            ]
        )
    set_excel_style(ws2)

    wb.save(workbook_path)

    grouped: dict[str, list[MultiFrameResult]] = {}
    for result in results:
        grouped.setdefault(result.variant, []).append(result)

    summary_rows: list[dict[str, object]] = []
    for variant, variant_results in grouped.items():
        successful = [result for result in variant_results if result.ok]
        latencies = sorted(result.latency_sec for result in successful)
        comments_generated = sum(
            len(result.comments) + len(result.long_comments)
            for result in successful
        )

        summary_rows.append(
            {
                "variant": variant,
                "use_multi_frame": variant_results[0].use_multi_frame,
                "history_count_requested": (
                    variant_results[0].history_count_requested
                ),
                "history_count_sent_avg": (
                    round(
                        statistics.fmean(
                            result.history_count_sent for result in successful
                        ),
                        3,
                    )
                    if successful
                    else ""
                ),
                "current_dimension": variant_results[0].current_dimension,
                "current_quality": variant_results[0].current_quality,
                "history_dimension": variant_results[0].history_dimension,
                "history_quality": variant_results[0].history_quality,
                "frames_expected": image_count,
                "frames_attempted": len(variant_results),
                "successes": len(successful),
                "failures": len(variant_results) - len(successful),
                "success_rate": round(
                    len(successful) / len(variant_results),
                    3,
                ),
                "comments_generated": comments_generated,
                "avg_comments_per_success": (
                    round(comments_generated / len(successful), 3)
                    if successful
                    else ""
                ),
                "avg_latency_sec": (
                    round(statistics.fmean(latencies), 3)
                    if latencies
                    else ""
                ),
                "median_latency_sec": (
                    round(statistics.median(latencies), 3)
                    if latencies
                    else ""
                ),
                "p90_latency_sec": (
                    round(percentile(latencies, 0.9), 3)
                    if latencies
                    else ""
                ),
                "min_latency_sec": min(latencies) if latencies else "",
                "max_latency_sec": max(latencies) if latencies else "",
                "avg_image_preparation_sec": (
                    round(
                        statistics.fmean(
                            result.image_preparation_sec for result in successful
                        ),
                        6,
                    )
                    if successful
                    else ""
                ),
                "avg_api_duration_sec": (
                    round(
                        statistics.fmean(
                            result.api_duration_sec for result in successful
                        ),
                        3,
                    )
                    if successful
                    else ""
                ),
                "avg_end_to_end_sec": (
                    round(
                        statistics.fmean(
                            result.end_to_end_sec for result in successful
                        ),
                        3,
                    )
                    if successful
                    else ""
                ),
                "retry_used_count": sum(
                    result.retry_used for result in variant_results
                ),
                "max_output_tokens": max_output_tokens,
            }
        )

    with summary_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=list(summary_rows[0].keys()),
        )
        writer.writeheader()
        writer.writerows(summary_rows)

    readme = (
        "Gemini multi-frame sequence benchmark\n\n"
        f"Model: {results[0].model}\n"
        f"Frames per variant: {image_count}\n"
        f"Variants: {', '.join(grouped)}\n"
        f"Maximum output tokens: {max_output_tokens}\n\n"
        "Single-frame and multi-frame variants process the same ordered "
        "screenshot sequence with isolated rolling summary and recent-comment "
        "history. The comparison outputs are intended for manual comment "
        "quality review as well as latency analysis.\n"
    )
    (output_dir / "README.txt").write_text(readme, encoding="utf-8")

    print(f"[done] wrote {frame_results_path}")
    print(f"[done] wrote {comparison_path}")
    print(f"[done] wrote {quality_review_path}")
    print(f"[done] wrote {summary_path}")
    print(f"[done] wrote {workbook_path}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Replay an ordered screenshot sequence through Gemini and compare "
            "single-frame versus multi-frame behavior."
        )
    )
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument(
        "--image-dir",
        default="logs/latency_benchmarks/for_benchmark",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Maximum ordered images per variant. Use 0 for all.",
    )
    parser.add_argument("--max-output-tokens", type=int, default=512)
    parser.add_argument("--retry-count", type=int, default=1)
    parser.add_argument(
        "--output-dir",
        default="logs/latency_benchmarks",
    )

    parser.add_argument("--single-dimension", type=int, default=768)
    parser.add_argument("--single-quality", type=int, default=72)
    parser.add_argument("--multi-current-dimension", type=int, default=768)
    parser.add_argument("--multi-current-quality", type=int, default=72)
    parser.add_argument("--history-count", type=int, default=3)
    parser.add_argument("--history-dimension", type=int, default=384)
    parser.add_argument("--history-quality", type=int, default=42)
    return parser


def main() -> int:
    if load_dotenv is not None:
        load_dotenv(REPO_ROOT / ".env")

    args = build_parser().parse_args()
    images = collect_images(Path(args.image_dir), args.limit)

    run_id = datetime.now().strftime("multiframe_%Y%m%d_%H%M%S")
    output_dir = Path(args.output_dir) / run_id
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[MultiFrameResult] = []

    variants = [
        VariantConfig(
            name="single_frame",
            current_dimension=args.single_dimension,
            current_quality=args.single_quality,
            use_multi_frame=False,
            history_count=0,
            history_dimension=0,
            history_quality=0,
        ),
        VariantConfig(
            name="multi_frame",
            current_dimension=args.multi_current_dimension,
            current_quality=args.multi_current_quality,
            use_multi_frame=True,
            history_count=args.history_count,
            history_dimension=args.history_dimension,
            history_quality=args.history_quality,
        ),
    ]

    print(
        f"[start] model={args.model} images={len(images)} "
        f"single={args.single_dimension}px/{args.single_quality} "
        f"multi={args.multi_current_dimension}px/{args.multi_current_quality} "
        f"history={args.history_count}x {args.history_dimension}px/{args.history_quality}"
    )
    print(
        f"[start] max_output_tokens={args.max_output_tokens} "
        f"retry_count={args.retry_count}"
    )

    for config in variants:
        client = build_client(
            model=args.model,
            current_dimension=config.current_dimension,
            current_quality=config.current_quality,
            max_output_tokens=args.max_output_tokens,
            history_dimension=config.history_dimension,
            history_quality=config.history_quality,
        )
        if client is None:
            return 1

        context = ModelContext()
        variant_dir = output_dir / "variants" / config.name
        print(f"[variant] starting {config.name}")

        for frame_index, image_path in enumerate(images, start=1):
            print(
                f"[run] {config.name} frame {frame_index}/{len(images)} "
                f"{image_path.name}"
            )
            result = generate_frame(
                client=client,
                model=args.model,
                config=config,
                image_path=image_path,
                frame_index=frame_index,
                images=images,
                context=context,
                retry_count=max(0, args.retry_count),
            )
            results.append(result)
            write_variant_log(
                result=result,
                variant_dir=variant_dir,
                max_output_tokens=args.max_output_tokens,
            )
            status = "ok" if result.ok else "error"
            print(
                f"[result] {config.name} {status} "
                f"history_sent={result.history_count_sent} "
                f"full={result.latency_sec}s "
                f"api={result.api_duration_sec}s "
                f"comments={len(result.comments) + len(result.long_comments)}"
            )

        print(f"[variant] finished {config.name}")

    write_outputs(
        results=results,
        output_dir=output_dir,
        image_count=len(images),
        max_output_tokens=args.max_output_tokens,
    )
    print(f"[done] multi-frame benchmark: {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
